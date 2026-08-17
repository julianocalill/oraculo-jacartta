# Preenche o bloco ANÁLISE JULIANO na planilha Análise Preço-Produto.
# Versão consolidada das regras finais de 16/08/2026:
#   - de-para por pedidos casados (01) > SKU do de-para > SKU idêntico > herança
#     entre variações do mesmo anúncio ESCALADA pela quantidade (30→120un = 4x)
#   - custo: anúncio de KIT (existe kit na Olist com aquela composição) → valor
#     da aba de kits (= custo médio do componente, ao vivo); unitário → preço de
#     custo do cadastro (ao vivo); fallback: espelho de junho, marcado
#   - Lucro/Prejuízo: fórmula do Juliano com guarda (branco sem custo)
#   - linhas com QTD/custo manuais são MANTIDAS (nota ATENÇÃO se divergirem)
# Entradas: pairs-itemmodel.json, cache-shopee*.json (de-para por SKU),
#   olist-products.json (espelho), olist-live-costs.json (02), kits.json (03)
# Uso: python 04-preencher.py <planilha-fonte.xlsx> <saida.xlsx>
import json, re, sys, glob
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font
from copy import copy
import os

S = os.environ.get("ANALISE_DIR", ".")
SRC, OUT = sys.argv[1], sys.argv[2]
norm = lambda s: str(s).strip().upper() if s not in (None, "") else None

def qnum(s):
    if s in (None, ""): return None
    m = re.search(r"(\d+)\s*(?:un\b|unid|unidades|potes?|p(?:ç|c)s?\b|pe(?:ç|c)as|rolos?|cabides?)",
                  str(s).lower())
    return int(m.group(1)) if m else None

pairs = json.load(open(f"{S}/pairs-itemmodel.json"))
by_pair = {}
for p in pairs:
    by_pair.setdefault((str(p["item_id"]), str(p["model_id"])), []).append(p)
for v in by_pair.values(): v.sort(key=lambda p: p["rk"])

cache = []
for f in sorted(glob.glob(f"{S}/cache-shopee*.json")): cache += json.load(open(f))
sku_map = {norm(r["channel_sku"]): r for r in cache
           if r["pair_rank"] == 1 and r["match_status"] == "mapeado" and r["channel_sku"]}

snap = {}
for p in json.load(open(f"{S}/olist-products.json")):
    k = norm(p["sku"])
    if not k: continue
    score = ((p.get("active") and 1 or 0), ((p["preco_custo"] or 0) > 0 and 1 or 0))
    if k not in snap or score > snap[k][0]: snap[k] = (score, p)
snap = {k: v[1] for k, v in snap.items()}
live = {norm(k): v for k, v in json.load(open(f"{S}/olist-live-costs.json")).items()}

kits = json.load(open(f"{S}/kits.json"))
kitmap = {}
for k in kits:
    if len(k["comp"]) != 1: continue
    c = k["comp"][0]
    kitmap.setdefault((norm(c["produto"]["sku"]), int(round(float(c["quantidade"])))), k["kit_sku"])

def campos(sku):
    k = norm(sku)
    lv = live.get(k)
    if lv: return (lv["precoCusto"] or 0, lv["precoCustoMedio"] or 0, "hoje")
    p = snap.get(k)
    if p: return (p["preco_custo"] or 0, p["preco_custo_medio"] or 0, "junho")
    return (0, 0, None)

def custo(sku, qtd):
    """(custo_unitario, tag) segundo a regra kit/unitário"""
    cad, medio, fonte = campos(sku)
    tag = " · custo de JUNHO (não conferido hoje)" if fonte == "junho" else ""
    kit_sku = kitmap.get((norm(sku), qtd))
    if kit_sku and medio > 0:
        return round(float(medio), 2), f" · custo do KIT {kit_sku} ({qtd}× {sku})" + tag
    if cad > 0:
        extra = (f" · ATENÇÃO: preço de custo {cad:.2f} vs médio {medio:.2f} — conferir cadastro"
                 if medio > 0 and cad > 3 * medio else "")
        return round(float(cad), 2), extra + tag
    if medio > 0:
        return round(float(medio), 2), " · sem preço de custo no cadastro; usei o médio" + tag
    return None, " · CUSTO ZERADO na Olist" + tag

wb = openpyxl.load_workbook(SRC)
ws = wb["Produtos"]
HDR = 4
rows = [r for r in range(HDR + 1, ws.max_row + 1) if ws.cell(r, 5).value not in (None, "")]

res = {}
for r in rows:
    E, F, D = str(ws.cell(r, 5).value), str(ws.cell(r, 6).value), ws.cell(r, 4).value
    entry = None
    plist = by_pair.get((E, F))
    if plist:
        p1 = plist[0]
        share = p1["orders_matched"] / p1["orders_total"]
        if share >= 0.8:
            entry = dict(sku=p1["sku_olist"], qtd=max(1, round(float(p1["qty_ratio"] or 1))),
                         origem=f"venda casada ({p1['orders_matched']} pedido{'s' if p1['orders_matched']>1 else ''})",
                         pedidos=p1["orders_matched"])
        else:
            alt = plist[1]["sku_olist"] if len(plist) > 1 else "?"
            res[r] = dict(sku=None, origem=f"ambíguo: {p1['sku_olist']} ({share:.0%}) vs {alt}",
                          pedidos=p1["orders_total"])
            continue
    if entry is None and norm(D) in sku_map:
        m = sku_map[norm(D)]
        entry = dict(sku=m["sku_olist"], qtd=max(1, round(float(m["qty_ratio"] or 1))),
                     origem="SKU do de-para", pedidos=m["orders_matched"])
    if entry is None and (norm(D) in snap or norm(D) in live):
        entry = dict(sku=(snap.get(norm(D)) or {}).get("sku") or D, qtd=1,
                     origem="SKU idêntico na Olist", pedidos=0)
    res[r] = entry or dict(sku=None, origem=None, pedidos=0)

by_item = defaultdict(list)
for r in rows: by_item[str(ws.cell(r, 5).value)].append(r)
for item, rs in by_item.items():
    solved = [r for r in rs if res[r].get("sku")]
    if not solved: continue
    if len(set(res[r]["sku"] for r in solved)) > 1: continue
    base = solved[0]
    sku_b, qtd_b = res[base]["sku"], res[base].get("qtd", 1)
    n_b = qnum(ws.cell(base, 3).value)
    for r in rs:
        if res[r].get("sku") or (res[r].get("origem") or "").startswith("ambíguo"): continue
        n_t = qnum(ws.cell(r, 3).value)
        if n_b and n_t:
            fator = qtd_b * n_t / n_b
            if abs(fator - round(fator)) < 0.01 and fator >= 1:
                res[r] = dict(sku=sku_b, qtd=int(round(fator)),
                              origem=f"herdado do anúncio (escalado {n_b}→{n_t} un)", pedidos=0)
            else:
                res[r] = dict(sku=None, origem=f"variação com quantidade própria ({n_t} un) — revisar", pedidos=0)
        elif n_b == n_t:
            res[r] = dict(sku=sku_b, qtd=qtd_b, origem="herdado do anúncio", pedidos=0)
        else:
            res[r] = dict(sku=None, origem="variação com quantidade própria — revisar", pedidos=0)

stats = Counter()
fmt_int, fmt_money = "0", "#,##0.00"
for r in rows:
    e = res[r]
    vM, vN, vO = ws.cell(r, 13).value, ws.cell(r, 14).value, ws.cell(r, 15).value
    manual = vM not in (None, "") or vN not in (None, "")
    sku = e.get("sku"); origem = e.get("origem")
    cu = tag = None
    if sku:
        cu, tag = custo(sku, e.get("qtd", 1))
        if tag: origem += tag
    if manual:
        origem = (origem + " · " if origem else "") + "QTD/custo manuais mantidos"
        if cu is not None and isinstance(vN, (int, float)) and abs(cu - vN) > 0.01:
            origem += f" · ATENÇÃO: Olist hoje diz {cu:.2f}"
        stats["manual mantido"] += 1
    elif sku and cu is not None:
        ws.cell(r, 13, e["qtd"]).number_format = fmt_int
        ws.cell(r, 14, cu).number_format = fmt_money
        c = ws.cell(r, 15); c.value = f"=ROUND(M{r}*N{r},2)"; c.number_format = fmt_money
        stats[e["origem"].split(" (")[0]] += 1
    else:
        if isinstance(vO, (int, float)) and vO == 0:
            ws.cell(r, 15).value = None  # zero envenena o lucro
            stats["custo 0 limpo"] += 1
        elif isinstance(vO, (int, float)) and vO > 0:
            origem = (origem + " · " if origem else "") + "custo preexistente (busca por texto) mantido"
            stats["preexistente mantido"] += 1
        else:
            stats["sem custo"] += 1
        if not origem: origem = "sem correspondência"
    f = (f'=IF($O{r}="","",L{r}-O{r}-(L{r}*IF(L{r}<=79.99,20%,14%)'
         f'+IF(L{r}<=79.99,4,IF(L{r}<=99.99,16,IF(L{r}<=199.99,20,IF(L{r}<=499.99,26,28)))))'
         f'-L{r}*1.3%-L{r}*6%-(L{r}-O{r})*9.25%-L{r}*3%-L{r}*3%-1)')
    cP = ws.cell(r, 16); cP.value = f; cP.number_format = fmt_money
    if ws.cell(r, 17).value in (None, ""):
        ws.cell(r, 17, f'=IF(AND(K{r}<>"",H{r}>0),(H{r}-K{r})/H{r},"")')
    ws.cell(r, 22, sku or "")
    ws.cell(r, 23, origem or "sem correspondência")
    ws.cell(r, 24, e.get("pedidos") or None)

for col, t in ((22, "SKU Olist"), (23, "Origem do custo"), (24, "Pedidos casados")):
    h = ws.cell(HDR, col, t); ref = ws.cell(HDR, 21)
    h.font, h.fill, h.border, h.alignment = copy(ref.font), copy(ref.fill), copy(ref.border), copy(ref.alignment)
ws.cell(3, 22, "PREENCHIDO PELO ORÁCULO (custos ao vivo na API da Olist)").font = Font(bold=True, size=9)
ws.cell(2, 23).value = ("Custo: anúncio de KIT usa o valor da aba de kits da Olist (componente × quantidade); "
                        "produto unitário usa o preço de custo do cadastro.")
ws.cell(2, 23).font = Font(size=8, bold=True)
for col, wd in ((22, 18), (23, 56), (24, 14)):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = wd

wb.save(OUT)
for k, v in stats.most_common(): print(f"  {k}: {v}")
