# Coluna Y "Checagem de modelo": compara dimensão (80x60 vs 60x60), volume (ml)
# e peso (kg) entre o texto do anúncio e o nome do produto Olist mapeado, e
# marca evidência fraca (1 pedido). Filtrar por "⚠" antes de precificar.
# Uso: python 06-checagem-modelo.py <planilha.xlsx>
import json, re, sys, os
import openpyxl
from copy import copy

S = os.environ.get("ANALISE_DIR", ".")
ARQ = sys.argv[1]
live = json.load(open(f"{S}/olist-live-costs.json"))
snap = {p["sku"]: p for p in json.load(open(f"{S}/olist-products.json")) if p["sku"]}

def nome_olist(sku):
    d = live.get(str(sku))
    if d and d.get("nome"): return str(d["nome"])
    p = snap.get(str(sku))
    return str(p["nome"]) if p else ""

def dims(txt):
    out = set()
    for m in re.finditer(r"(\d{2,3})\s*[xX×]\s*(\d{2,3})", str(txt)):
        a, b = sorted((int(m.group(1)), int(m.group(2))), reverse=True)
        out.add(f"{a}x{b}")
    return out
vols = lambda t: set(re.findall(r"(\d{3,4})\s*ml", str(t).lower()))
pesos = lambda t: set(re.findall(r"(\d{1,2}(?:[.,]\d)?)\s*kg", str(t).lower()))

wb = openpyxl.load_workbook(ARQ)
ws = wb["Produtos"]
COL = 25
conflitos = 0
for r in range(5, ws.max_row + 1):
    if ws.cell(r, 5).value in (None, ""): continue
    sku = ws.cell(r, 22).value
    flags = []
    if sku not in (None, ""):
        ad = f"{ws.cell(r,2).value or ''} {ws.cell(r,3).value or ''}"
        ol = nome_olist(sku)
        d_ad, d_ol = dims(ad), dims(ol)
        if d_ad and d_ol and not (d_ad & d_ol):
            flags.append(f"MODELO: anúncio diz {'/'.join(sorted(d_ad))}, Olist é {'/'.join(sorted(d_ol))}")
        v_ad, v_ol = vols(ad), vols(ol)
        if v_ad and v_ol and not (v_ad & v_ol):
            flags.append(f"VOLUME: anúncio {'/'.join(sorted(v_ad))}ml vs Olist {'/'.join(sorted(v_ol))}ml")
        p_ad, p_ol = pesos(ad), pesos(ol)
        if p_ad and p_ol and not (p_ad & p_ol):
            flags.append(f"PESO: anúncio {'/'.join(sorted(p_ad))}kg vs Olist {'/'.join(sorted(p_ol))}kg")
        if re.search(r"venda casada \(1 pedido\)", str(ws.cell(r, 23).value or "")):
            flags.append("evidência fraca (1 pedido)")
    if flags:
        ws.cell(r, COL, "⚠ " + " · ".join(flags))
        if any(f.startswith(("MODELO", "VOLUME", "PESO")) for f in flags): conflitos += 1
    else:
        ws.cell(r, COL, "ok" if sku not in (None, "") else "")
h = ws.cell(4, COL, "Checagem de modelo")
ref = ws.cell(4, 21)
h.font, h.fill, h.border, h.alignment = copy(ref.font), copy(ref.fill), copy(ref.border), copy(ref.alignment)
ws.column_dimensions["Y"].width = 46
wb.save(ARQ)
print("conflitos de modelo:", conflitos)
