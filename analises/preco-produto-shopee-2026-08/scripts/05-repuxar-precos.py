# Repuxa os preços atuais (colunas I e K) do sync Shopee do Oráculo
# (shopee_products, sincronizado de hora em hora). L e P recalculam sozinhos.
# Antes: baixar shopee_products via REST paginado (item_id, model_id,
# model_price, price_min) para shopee-prices-*.json — ver README.
# Uso: python 05-repuxar-precos.py <planilha.xlsx>
import json, glob, sys, os
import openpyxl
from openpyxl.styles import Font

S = os.environ.get("ANALISE_DIR", ".")
ARQ = sys.argv[1]
precos = {}
for f in sorted(glob.glob(f"{S}/shopee-prices-*.json")):
    for p in json.load(open(f)):
        key = (str(p["item_id"]), str(p["model_id"] if p["model_id"] not in (None, "") else "0"))
        v = p["model_price"] if p["model_price"] is not None else p["price_min"]
        if v is not None: precos[key] = float(v)

wb = openpyxl.load_workbook(ARQ)
ws = wb["Produtos"]
FLAG = "preço não conferido (anúncio fora do sync de hoje)"
mudou = nao_achado = 0
for r in range(5, ws.max_row + 1):
    if ws.cell(r, 5).value in (None, ""): continue
    key = (str(ws.cell(r, 5).value), str(ws.cell(r, 6).value))
    novo = precos.get(key)
    w = ws.cell(r, 23)
    if novo is None:
        nao_achado += 1
        if FLAG not in str(w.value or ""):
            w.value = (str(w.value) + " · " if w.value else "") + FLAG
        continue
    promo = str(ws.cell(r, 10).value or "").upper() == "SIM"
    antigo = ws.cell(r, 11).value if promo else ws.cell(r, 9).value
    ws.cell(r, 9, novo)
    if promo: ws.cell(r, 11, novo)
    if isinstance(antigo, (int, float)) and abs(novo - antigo) > 0.005: mudou += 1
wb.save(ARQ)
print(f"preço mudou: {mudou} | não achado: {nao_achado}")
